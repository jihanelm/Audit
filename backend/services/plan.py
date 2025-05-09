import os
from copy import copy
from datetime import datetime, date
from io import BytesIO
from typing import Optional, List

import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
from fastapi import HTTPException, UploadFile
from openpyxl.styles import PatternFill
from sqlalchemy import extract, func
from sqlalchemy.orm import Session, joinedload
from backend.models.plan import Plan
from backend.models.vulnerability import Vulnerability
from backend.schemas.plan import PlanUpdate, VulnerabilitySummary, PlanResponse

from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.drawing.image import Image as XLImage

from openpyxl.styles import Font, Alignment

from backend.schemas.vulnerability import VulnerabiliteResponse
from log_config import setup_logger

logger = setup_logger()

def generate_plan_ref(session: Session, date_realisation: date) -> str:
    year = date_realisation.year
    total_existing = session.query(func.count()).filter(
        func.extract('year', Plan.date_realisation) == year
    ).scalar() or 0

    letter_index = total_existing // 99
    if letter_index >= 26:
        raise ValueError("Trop de plans pour l'année !")

    letter = chr(ord('A') + letter_index)
    number = (total_existing % 99) + 1
    return f"{year}_{letter}_{number:02d}"

async def process_uploaded_plan(file: UploadFile, db: Session):
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Format de fichier non supporté.")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        df.replace({np.nan: None}, inplace=True)

        required_columns = {
            "ref", "application", "type_application", "type_audit",
            "date_realisation", "date_cloture", "date_rapport",
            "nb_vulnerabilites", "niveau_securite", "commentaire_dcsg",
            "commentaire_cp", "taux_remediation",
            "titre", "criticite", "pourcentage_remediation", "statut_remediation", "actions"
        }

        if not required_columns.issubset(df.columns):
            raise HTTPException(status_code=400, detail=f"Colonnes manquantes : {required_columns - set(df.columns)}")

        # Nettoyage
        df = df.dropna(subset=['ref', 'date_realisation'])

        grouped = defaultdict(list)
        for index, row in df.iterrows():
            grouped[row["ref"]].append(row.to_dict())

        for ref, rows in grouped.items():
            first_row = rows[0]
            try:
                date_realisation = pd.to_datetime(first_row["date_realisation"], dayfirst=True, errors="coerce")
                if pd.isnull(date_realisation):
                    raise ValueError("Date de réalisation invalide.")

                plan_ref = generate_plan_ref(db, date_realisation)

                plan = Plan(
                    ref=plan_ref,
                    application=first_row["application"],
                    type_application=first_row["type_application"],
                    type_audit=first_row["type_audit"],
                    date_realisation=date_realisation,
                    date_cloture=pd.to_datetime(first_row.get("date_cloture")) if first_row.get("date_cloture") else None,
                    date_rapport=pd.to_datetime(first_row.get("date_rapport")) if first_row.get("date_rapport") else None,
                    niveau_securite=first_row.get("niveau_securite"),
                    taux_remediation=first_row.get("taux_remediation"),
                    commentaire_dcsg=first_row.get("commentaire_dcsg"),
                    commentaire_cp=first_row.get("commentaire_cp")
                )

                db.add(plan)
                db.flush()
                db.refresh(plan)

                for row in rows:
                    vuln = Vulnerability(
                        plan_id=plan.id,
                        titre=row.get("titre"),
                        criticite=row.get("criticite"),
                        pourcentage_remediation=row.get("pourcentage_remediation"),
                        statut_remediation=row.get("statut_remediation"),
                        actions=row.get("actions")
                    )
                    db.add(vuln)

                db.flush()
                vulnerabilites = db.query(Vulnerability).filter_by(plan_id=plan.id).all()
                summary = compute_vulnerability_summary(vulnerabilites)
                plan.nb_vulnerabilites = VulnerabilitySummary(**compute_vulnerability_summary(vulnerabilites)).dict()
                plan.taux_remediation = compute_taux_remediation(vulnerabilites)

                db.commit()

            except Exception as e:
                logger.error(f"Erreur insertion vulnérabilité pour ref {ref} : {e}")
                db.rollback()

        logger.info("Plans et vulnérabilités insérés avec succès.")
        return {"message": "Importation réussie"}

    except Exception as e:
        error_message = str(e.orig) if hasattr(e, 'orig') else str(e)
        logger.error(f"Erreur lors du traitement du fichier : {error_message}")
        raise HTTPException(status_code=500, detail="Erreur de traitement du fichier.")

def export_plans_to_excel(
    db: Session,
    ref: Optional[str] = None,
    application: Optional[str] = None,
    type_audit: Optional[str] = None,
    niveau_securite: Optional[str] = None,
    date_realisation: Optional[str] = None,
    date_cloture: Optional[str] = None,
    date_rapport: Optional[str] = None,
    realisation_year: Optional[int] = None,
    realisation_month: Optional[int] = None,
    cloture_year: Optional[int] = None,
    cloture_month: Optional[int] = None,
    rapport_year: Optional[int] = None,
    rapport_month: Optional[int] = None,
):
    query = db.query(Plan).options(joinedload(Plan.vulnerabilites))

    # Filtres
    if ref:
        query = query.filter(Plan.ref.ilike(f"%{ref}%"))
    if application:
        query = query.filter(Plan.application.ilike(f"%{application}%"))
    if type_audit:
        query = query.filter(Plan.type_audit == type_audit)
    if niveau_securite:
        query = query.filter(Plan.niveau_securite == niveau_securite)
    if date_realisation:
        query = query.filter(Plan.date_realisation == date_realisation)
    if date_cloture:
        query = query.filter(Plan.date_cloture == date_cloture)
    if date_rapport:
        query = query.filter(Plan.date_rapport == date_rapport)
    if realisation_year:
        query = query.filter(extract('year', Plan.date_realisation) == realisation_year)
    if realisation_month:
        query = query.filter(extract('month', Plan.date_realisation) == realisation_month)
    if cloture_year:
        query = query.filter(extract('year', Plan.date_cloture) == cloture_year)
    if cloture_month:
        query = query.filter(extract('month', Plan.date_cloture) == cloture_month)
    if rapport_year:
        query = query.filter(extract('year', Plan.date_rapport) == rapport_year)
    if rapport_month:
        query = query.filter(extract('month', Plan.date_rapport) == rapport_month)

    plans = query.all()
    if not plans:
        return None

    # Construction des données
    combined_data = []
    for plan in plans:
        if plan.vulnerabilites:
            for vuln in plan.vulnerabilites:
                combined_data.append({
                    "Réf": plan.ref,
                    "Application/Solution": plan.application,
                    "Type d'application": plan.type_application,
                    "Type d'audit": plan.type_audit,
                    "Date de realisation de la mission": plan.date_realisation,
                    "Date de cloture de la mission": plan.date_cloture,
                    "Date de communication du rapport": plan.date_rapport,
                    "Niveau de securité": plan.niveau_securite,
                    "Nombre de vulnérabilités": format_vulnerabilites(plan.nb_vulnerabilites),
                    "Commentaire DCSG": clean_html(plan.commentaire_dcsg),
                    "Commentaire CP": clean_html(plan.commentaire_cp),
                    "Titre vulnérabilité": vuln.titre,
                    "Criticité": vuln.criticite,
                    "Pourcentage de remédiation": vuln.pourcentage_remediation,
                    "Statut de remédiation": vuln.statut_remediation,
                    "Actions": vuln.actions
                })
        else:
            combined_data.append({
                "Réf": plan.ref,
                "Application/Solution": plan.application,
                "Type d'application": plan.type_application,
                "Type d'audit": plan.type_audit,
                "Date de realisation de la mission": plan.date_realisation,
                "Date de cloture de la mission": plan.date_cloture,
                "Date de communication du rapport": plan.date_rapport,
                "Niveau de securité": plan.niveau_securite,
                "Nombre de vulnérabilités": plan.nb_vulnerabilites,
                "Commentaire DCSG": plan.commentaire_dcsg,
                "Commentaire CP": plan.commentaire_cp,
                "Titre vulnérabilité": "",
                "Criticité": "",
                "Pourcentage de remédiation": "",
                "Statut de remédiation": "",
                "Actions": ""
            })

    # Chemin vers le modèle de la page de garde
    current_dir = os.path.dirname(os.path.abspath(__file__))
    cover_path = os.path.join(current_dir, '..', '..', 'Page_de_garde.xlsx')

    # Charger la page de garde existante
    cover_wb = load_workbook(cover_path)
    cover_ws = cover_wb.active

    # Nouveau fichier final
    final_wb = Workbook()
    final_wb.remove(final_wb.active)

    CRITICITE_COLOR_MAP = {
        "mineure": PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid"),
        "moderee": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "majeure": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "critique": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    }

    # Copier chaque cellule avec son style
    new_cover_ws = final_wb.create_sheet("Page de garde")
    for row in cover_ws.iter_rows():
        for cell in row:
            new_cell = new_cover_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Gérer la largeur des colonnes
    for col_letter, dim in cover_ws.column_dimensions.items():
        new_cover_ws.column_dimensions[col_letter].width = dim.width

    # Gérer les fusions
    for merged_cell in cover_ws.merged_cells.ranges:
        new_cover_ws.merge_cells(str(merged_cell))

    # Ajouter une image à la page de garde (entre C18 et F24)
    image_path = os.path.join(current_dir, '..', '..', 'pictures', 'logo.png')
    if os.path.exists(image_path):
        img = XLImage(image_path)

        img.width = 64 * 4  # ≈ 256 pixels
        img.height = 20 * 7  # ≈ 140 pixels

        new_cover_ws.add_image(img, 'C18')  # Position d’ancrage

    # Ajouter les données "Plans"
    plans_ws = final_wb.create_sheet("Plans")
    df = pd.DataFrame(combined_data)

    for r in dataframe_to_rows(df, index=False, header=True):
        plans_ws.append(r)

    # Appliquer une mise en forme à la première ligne (en-tête)
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in plans_ws[1]:  # Ligne 1 = en-tête
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Appliquer le style de couleur selon la criticité
    criticite_column_index = None
    for idx, cell in enumerate(plans_ws[1], start=1):
        if cell.value == "Criticité":
            criticite_column_index = idx
            break

    if criticite_column_index:
        for row in plans_ws.iter_rows(min_row=2, min_col=criticite_column_index, max_col=criticite_column_index):
            for cell in row:
                criticite_value = str(cell.value).strip().lower() if cell.value else ""
                if criticite_value in CRITICITE_COLOR_MAP:
                    cell.fill = CRITICITE_COLOR_MAP[criticite_value]

    # Sauvegarde
    EXPORT_DIR = os.path.join(current_dir, '..', '..', 'exports_plan')
    os.makedirs(EXPORT_DIR, exist_ok=True)
    file_path = os.path.join(EXPORT_DIR, f"plans_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    final_wb.save(file_path)

    return file_path

def get_filtered_plans(
    db: Session,
    ref: Optional[str] = None,
    application: Optional[str] = None,
    type_audit: Optional[str] = None,
    niveau_securite: Optional[str] = None,
    # Filtres exacts
    date_realisation: Optional[str] = None,
    date_cloture: Optional[str] = None,
    date_rapport: Optional[str] = None,
    # Filtres par année/mois
    realisation_year: Optional[int] = None,
    realisation_month: Optional[int] = None,
    cloture_year: Optional[int] = None,
    cloture_month: Optional[int] = None,
    rapport_year: Optional[int] = None,
    rapport_month: Optional[int] = None,
):
    query = db.query(Plan).options(joinedload(Plan.vulnerabilites))

    # Filtres textuels
    if ref:
        query = query.filter(Plan.ref.ilike(f"%{ref}%"))
    if application:
        query = query.filter(Plan.application.ilike(f"%{application}%"))
    if type_audit:
        query = query.filter(Plan.type_audit == type_audit)
    if niveau_securite:
        query = query.filter(Plan.niveau_securite == niveau_securite)

    # Filtres exacts sur les dates
    if date_realisation:
        query = query.filter(Plan.date_realisation == date_realisation)
    if date_cloture:
        query = query.filter(Plan.date_cloture == date_cloture)
    if date_rapport:
        query = query.filter(Plan.date_rapport == date_rapport)

    # Filtres par année/mois
    if realisation_year:
        query = query.filter(extract('year', Plan.date_realisation) == realisation_year)
    if realisation_month:
        query = query.filter(extract('month', Plan.date_realisation) == realisation_month)

    if cloture_year:
        query = query.filter(extract('year', Plan.date_cloture) == cloture_year)
    if cloture_month:
        query = query.filter(extract('month', Plan.date_cloture) == cloture_month)

    if rapport_year:
        query = query.filter(extract('year', Plan.date_rapport) == rapport_year)
    if rapport_month:
        query = query.filter(extract('month', Plan.date_rapport) == rapport_month)

    return query.all()

def update_plan(db: Session, plan_id: int, updated_data: PlanUpdate, vulnerabilites=None):
    plan = db.query(Plan).filter(Plan.id == plan_id).first()

    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé.")

    try:
        update_fields = updated_data.dict(exclude_unset=True)
        vulnerabilites_data = update_fields.pop("vulnerabilites", None)

        # Mise à jour des champs du plan
        for key, value in update_fields.items():
            setattr(plan, key, value)

        # Si les vulnérabilités sont incluses
        if vulnerabilites_data is not None:
            # Supprimer les vulnérabilités existantes
            db.query(Vulnerability).filter(Vulnerability.plan_id == plan.id).delete()

            # Ajouter les nouvelles
            new_vulns = []
            for vuln_data in vulnerabilites_data:
                vuln = Vulnerability(plan_id=plan.id, **vuln_data)
                db.add(vuln)
                new_vulns.append(vuln)

            db.flush()  # s'assurer que les vulnérabilités sont bien insérées avant résumé

            # Mettre à jour les champs résumés à partir des nouvelles vulnérabilités
            summary = compute_vulnerability_summary(new_vulns)
            plan.nb_vulnerabilites = dict(summary)
            plan.taux_remediation = compute_taux_remediation(new_vulns)

        db.commit()
        db.refresh(plan)

        logger.info(f"Plan {plan_id} et vulnérabilités mis à jour avec succès.")
        return plan

    except Exception as e:
        db.rollback()
        error_message = str(e.orig) if hasattr(e, 'orig') else str(e)
        logger.error(f"Erreur lors de la mise à jour du plan : {error_message}")
        raise HTTPException(status_code=500, detail="Erreur lors de la mise à jour du plan.")

def serialize_plan(plan: Plan) -> PlanResponse:
    return PlanResponse(
        id=plan.id,
        ref=plan.ref,
        application=plan.application,
        type_application=plan.type_application,
        type_audit=plan.type_audit,
        date_realisation=plan.date_realisation,
        date_cloture=plan.date_cloture,
        date_rapport=plan.date_rapport,
        niveau_securite=plan.niveau_securite,
        taux_remediation=plan.taux_remediation,
        commentaire_dcsg=plan.commentaire_dcsg,
        commentaire_cp=plan.commentaire_cp,
        nb_vulnerabilites=VulnerabilitySummary(**plan.nb_vulnerabilites)
        if isinstance(plan.nb_vulnerabilites, dict) else None,

        vulnerabilites=[
            VulnerabiliteResponse(
                id=v.id,
                titre=v.titre,
                criticite=v.criticite,
                pourcentage_remediation=v.pourcentage_remediation,
                statut_remediation=v.statut_remediation,
                actions=v.actions
            ) for v in plan.vulnerabilites
        ]
    )

def compute_vulnerability_summary(vulnerabilities):
    criticity_counts = Counter([v.criticite for v in vulnerabilities if v.criticite])
    return {
        "critique": criticity_counts.get("critique", 0),
        "majeure": criticity_counts.get("majeure", 0),
        "moderee": criticity_counts.get("moderee", 0),
        "mineure": criticity_counts.get("mineure", 0),
        "total": len(vulnerabilities)
    }

def compute_taux_remediation(vulnerabilites: List[Vulnerability]) -> float:
    valeurs = [v.pourcentage_remediation for v in vulnerabilites if v.pourcentage_remediation is not None]
    if not valeurs:
        return 0.0
    return round(sum(valeurs) / len(valeurs), 2)


def format_vulnerabilites(vuln_dict):
    if not vuln_dict:
        return "0 vulnérabilités"

    total = vuln_dict.get("total", 0)
    critique = vuln_dict.get("critique", 0)
    majeure = vuln_dict.get("majeure", 0)
    moderee = vuln_dict.get("moderee", 0)
    mineure = vuln_dict.get("mineure", 0)

    lines = [f"{total} vulnérabilités"]
    if critique:
        lines.append(f"({critique}) Critiques")
    if majeure:
        lines.append(f"({majeure}) Majeure")
    if moderee:
        lines.append(f"({moderee}) Moderee")
    if mineure:
        lines.append(f"({mineure}) Mineur")

    return "\n".join(lines)

def clean_html(raw_html):
    if not raw_html:
        return ""
    return BeautifulSoup(raw_html, "html.parser").get_text(separator=" ").strip()